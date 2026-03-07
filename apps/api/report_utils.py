import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from .models import ConsumptionLog, WeightRecord, WaterLog, DailyMealLog

def export_to_excel(user_profile, logs, weight_records, water_records, meal_logs, summary):
    """Exports health data to an Excel file buffer"""
    wb = Workbook()
    
    # Summary Sheet
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.append(["Health Report Summary for", user_profile.name])
    ws1.append(["Generated Date", ""]) 
    ws1.append([])
    ws1.append(["AI Progress Analysis:"])
    ws1.append([summary])
    ws1["A5"].alignment = Alignment(wrapText=True)
    
    # Logs Sheet
    ws2 = wb.create_sheet(title="Consumption Logs")
    ws2.append(["Date", "Type", "Details", "Calories (kcal)"])
    
    # Add regular logs
    for log in logs:
        ws2.append([str(log.date), log.meal_type, log.food_item.name, log.total_calories])
        
    # Add DailyMealLog entries
    for m in meal_logs:
        if m.breakfast_content:
            ws2.append([str(m.date), "Breakfast (Plan)", m.breakfast_content[:50], m.breakfast_calories])
        if m.lunch_content:
            ws2.append([str(m.date), "Lunch (Plan)", m.lunch_content[:50], m.lunch_calories])
        if m.dinner_content:
            ws2.append([str(m.date), "Dinner (Plan)", m.dinner_content[:50], m.dinner_calories])
        if m.snacks_content:
            ws2.append([str(m.date), "Snacks (Plan)", m.snacks_content[:50], m.snacks_calories])
    
    # Weight Sheet
    ws3 = wb.create_sheet(title="Weight & Water History")
    ws3.append(["Date", "Weight (kg)", "Water (Glasses)", "Hydration Goal Met"])
    
    # Create a map for dates
    date_data = {}
    for w in weight_records:
        date_data[w.date] = {'weight': w.weight, 'water': 0, 'met': 'No'}
    for wa in water_records:
        if wa.date not in date_data:
            date_data[wa.date] = {'weight': '--', 'water': wa.amount_glasses, 'met': 'Yes' if wa.is_target_completed else 'No'}
        else:
            date_data[wa.date]['water'] = wa.amount_glasses
            date_data[wa.date]['met'] = 'Yes' if wa.is_target_completed else 'No'

    for d in sorted(date_data.keys(), reverse=True):
        ws3.append([str(d), date_data[d]['weight'], date_data[d]['water'], date_data[d]['met']])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_to_pdf(user_profile, logs, weight_records, water_records, meal_logs, summary):
    """Exports health data to a PDF file buffer"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title = Paragraph(f"NutriDiet Health Report: {user_profile.name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary
    elements.append(Paragraph("AI Summary & Recommendations", styles['Heading2']))
    elements.append(Paragraph(summary, styles['Normal']))
    elements.append(Spacer(1, 24))
    
    # Weight & Hydration Table
    elements.append(Paragraph("Weight & Hydration History", styles['Heading3']))
    data = [["Date", "Weight", "Water", "Goal"]]
    
    date_data = {}
    for w in weight_records[:15]:
        date_data[w.date] = {'w': f"{w.weight}kg", 'wa': '--', 'st': '--'}
    for wa in water_records[:15]:
        if wa.date not in date_data:
            date_data[wa.date] = {'w': '--', 'wa': f"{wa.amount_glasses} gls", 'st': 'Done' if wa.is_target_completed else '--'}
        else:
            date_data[wa.date]['wa'] = f"{wa.amount_glasses} gls"
            date_data[wa.date]['st'] = 'Done' if wa.is_target_completed else '--'

    for d in sorted(date_data.keys(), reverse=True):
        data.append([str(d), date_data[d]['w'], date_data[d]['wa'], date_data[d]['st']])
        
    t = Table(data, colWidths=[100, 100, 100, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke])
    ]))
    elements.append(t)
    elements.append(Spacer(1, 24))
    
    # Consumption Table (Individual and Plan-based)
    elements.append(Paragraph("Food Consumption Logs", styles['Heading3']))
    cons_data = [["Date", "Type", "Details", "Calories"]]
    
    # Graphs Section
    if weight_records.exists() or water_records.exists():
        elements.append(Paragraph("Visual Data Trends", styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        # Weight Chart
        if weight_records.exists():
            plt.figure(figsize=(6, 3))
            w_dates = [w.date for w in weight_records.order_by('date')]
            w_values = [float(w.weight) for w in weight_records.order_by('date')]
            plt.plot(w_dates, w_values, marker='o', color='#3b82f6', linewidth=2)
            plt.fill_between(w_dates, w_values, alpha=0.1, color='#3b82f6')
            plt.title('Weight Trend (kg)', fontsize=10)
            plt.grid(True, alpha=0.2)
            plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
            plt.xticks(rotation=45, fontsize=8)
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150)
            plt.close()
            img_buffer.seek(0)
            elements.append(Image(img_buffer, width=400, height=200))
            elements.append(Spacer(1, 12))

        # Water Chart
        if water_records.exists():
            plt.figure(figsize=(6, 3))
            wa_dates = [w.date for w in water_records.order_by('date')]
            wa_values = [w.amount_glasses for w in water_records.order_by('date')]
            plt.bar(wa_dates, wa_values, color='#0ea5e9', alpha=0.7)
            plt.title('Water Intake (Glasses)', fontsize=10)
            plt.grid(axis='y', alpha=0.2)
            plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
            plt.xticks(rotation=45, fontsize=8)
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150)
            plt.close()
            img_buffer.seek(0)
            elements.append(Image(img_buffer, width=400, height=200))
            elements.append(Spacer(1, 24))

    # Final Footer
    combined_logs = []
    for l in logs:
        combined_logs.append({'d': l.date, 't': l.meal_type, 'de': l.food_item.name, 'c': f"{l.total_calories} kcal"})
    for m in meal_logs:
        if m.breakfast_content: combined_logs.append({'d': m.date, 't': 'Breakfast (P)', 'de': m.breakfast_content[:40], 'c': f"{m.breakfast_calories} kcal"})
        if m.lunch_content: combined_logs.append({'d': m.date, 't': 'Lunch (P)', 'de': m.lunch_content[:40], 'c': f"{m.lunch_calories} kcal"})
        if m.dinner_content: combined_logs.append({'d': m.date, 't': 'Dinner (P)', 'de': m.dinner_content[:40], 'c': f"{m.dinner_calories} kcal"})
        if m.snacks_content: combined_logs.append({'d': m.date, 't': 'Snacks (P)', 'de': m.snacks_content[:40], 'c': f"{m.snacks_calories} kcal"})
        
    for cl in sorted(combined_logs, key=lambda x: x['d'], reverse=True)[:20]:
        cons_data.append([str(cl['d']), cl['t'], cl['de'], cl['c']])
        
    ct = Table(cons_data, colWidths=[80, 80, 160, 80])
    ct.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(ct)
    elements.append(Spacer(1, 24))

    # Final Footer
    elements.append(Paragraph("End of Report - NutriDiet AI", styles['Italic']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer
